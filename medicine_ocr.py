"""
药品信息识别引擎
支持两种模式：
  1. 本地 EasyOCR 提取文字 + 正则解析
  2. OpenAI 兼容 API 视觉识别（更准确，可补全）
"""
import re
import os
import base64
from io import BytesIO
from PIL import Image
import json
from typing import Optional

# ===== 常量 =====
MEDICINE_FIELDS = ["药名", "厂家", "国药准字", "规格"]

# 英文/别名键 -> 中文标准键 的映射，兼容模型返回英文键名的情况
_FIELD_ALIASES = {
    "药名": ["药名", "药品名", "药品名称", "通用名", "名称", "name", "drug_name", "medicine_name"],
    "厂家": ["厂家", "生产厂家", "生产企业", "制造商", "公司", "manufacturer", "factory", "company", "producer"],
    "国药准字": ["国药准字", "批准文号", "批号", "国药准字h", "approval", "approval_no", "license", "registration_no"],
    "规格": ["规格", "spec", "specification", " packing", "package", "size"],
}

def _normalize_fields(item: dict) -> dict:
    """将一条识别结果统一为标准中文键名，缺失字段填空字符串。"""
    if not isinstance(item, dict):
        return {f: "" for f in MEDICINE_FIELDS}
    out = {f: "" for f in MEDICINE_FIELDS}
    for std, aliases in _FIELD_ALIASES.items():
        for alias in aliases:
            if alias in item and item[alias] not in (None, ""):
                out[std] = str(item[alias]).strip()
                break
    return out


RE_GUOYAO_FULL = re.compile(r"^国药准字[HZSBJ]\d{8}$")

def _validate_result(item: dict) -> dict:
    """对归一化后的结果做格式校验，异常记录到 _warn。"""
    warns = []
    gy = (item.get("国药准字") or "").replace(" ", "").replace("\u3000", "")
    if gy and not RE_GUOYAO_FULL.match(gy):
        warns.append("国药准字格式异常")
    if not (item.get("药名") or "").strip():
        warns.append("缺少药名")
    item["_warn"] = warns
    return item


# ===== EasyOCR 引擎（本地） =====
_easyocr_reader = None

def _get_reader(model_dir=None):
    global _easyocr_reader
    if _easyocr_reader is None:
        import easyocr
        kwargs = {"lang_list": ["ch_sim", "en"], "gpu": False}
        if model_dir:
            kwargs["model_storage_directory"] = model_dir
        _easyocr_reader = easyocr.Reader(**kwargs)
    return _easyocr_reader

def is_easyocr_model_ready(model_dir: str = None) -> bool:
    """检查本地 EasyOCR 模型是否已下载（首次使用需要约 50MB 模型文件）"""
    if model_dir is None:
        model_dir = os.path.join(os.path.expanduser("~"), ".EasyOCR", "model")
    if not os.path.isdir(model_dir):
        return False
    pth_files = [f for f in os.listdir(model_dir) if f.endswith(".pth")]
    # 过滤掉空文件 / 下载中断残留
    valid = [f for f in pth_files if os.path.getsize(os.path.join(model_dir, f)) > 1024]
    return len(valid) >= 1


def extract_text_easyocr(image_path: str, model_dir: str = None) -> str:
    """使用 EasyOCR 提取图片中的文字"""
    reader = _get_reader(model_dir)
    results = reader.readtext(image_path)
    lines = [r[1] for r in results]
    return "\n".join(lines)


# ===== 正则解析（EasyOCR 结果的后处理）=====
# 国药准字：国药准字H/Z/S/B + 8位数字
RE_GUOYAO = re.compile(r"国药准字\s*([HZSShzsb]?)\s*(\d{8})")
RE_GUOYAO_SIMPLE = re.compile(r"(?:国药准字|国药)?\s*[HZSShzsb]?\d{8}")

# 规格：常见模式
RE_GUIGE = re.compile(
    r"(?:规格[：:]?\s*)?"
    r"("
    r"\d+[\.,]?\d*\s*(?:mg|g|ml|ug|IU|单位|毫克|毫升|克)"
    r"(?:\*{1,2}\s*\d+\s*[粒片支袋板瓶盒]?)?"
    r"(?:\s*/\s*(?:板|盒|瓶|袋|支))?"
    r")",
    re.IGNORECASE
)

# 厂家：常见后缀
RE_CHANGJIA = re.compile(
    r"([\u4e00-\u9fffA-Za-z0-9（()）\u3001\u3002·]{4,}(?:有限公司|药业|制药|药厂|生物|科技|大药房|厂|集团))",
    re.IGNORECASE
)

def parse_medicine_info(raw_text: str) -> dict:
    """从 OCR 原始文本中解析药品信息
    支持两种布局：
    1. 标签式："药品名称: 绞股蓝总苷胶囊"
    2. 标签式换行：每行一个字段
    3. 纯文本排版：药名单独一行，厂家/准字号各自占行
    """
    info = {"药名": "", "厂家": "", "国药准字": "", "规格": ""}

    if not raw_text:
        return info

    original_text = raw_text
    # 保留换行，只去多余空格
    lines = [ln.strip() for ln in raw_text.splitlines() if ln.strip()]

    # 1. 国药准字：全文搜索
    m = RE_GUOYAO.search(raw_text)
    if m:
        cat = m.group(1).strip().upper()
        num = m.group(2).strip()
        info["国药准字"] = f"国药准字{cat}{num}"
    else:
        m2 = RE_GUOYAO_SIMPLE.search(raw_text)
        if m2:
            info["国药准字"] = m2.group(0).strip()

    # 2. 规格：全文搜索
    m = RE_GUIGE.search(raw_text)
    if m:
        info["规格"] = (m.group(1) or m.group(2) or "").strip()

    # 3. 厂家：先找关键词行，再用正则
    for ln in lines:
        if any(kw in ln for kw in ["生产企业", "生产厂家", "厂家", "制药", "药业"]) and len(ln) > 3:
            # 去掉关键词前缀
            cleaned = re.sub(r"^(生产企业|生产厂家|厂家)[:：\s]*", "", ln)
            if len(cleaned) >= 2:
                info["厂家"] = cleaned
                break
    if not info["厂家"]:
        m = RE_CHANGJIA.search(raw_text)
        if m:
            info["厂家"] = m.group(1).strip()

    # 4. 药名：多策略
    # 4.1 显式标签
    for kw in ["【药品名称】", "【通用名称】", "【通用名】", "药品名称", "通用名称", "通用名", "商品名", "产品名称"]:
        for ln in lines:
            if kw in ln:
                # 取冒号后或关键词后
                after = ln.split(kw, 1)[-1].strip(" :：")
                # 只取第一行，去掉英文/符号
                name = after.split()[0] if after else ""
                name = re.sub(r"[^\u4e00-\u9fffA-Za-z0-9]", "", name)
                if len(name) >= 2:
                    info["药名"] = name
                    break
        if info["药名"]:
            break

    # 4.2 启发式：找第一个像药名的行（中文+字母数字，2-15字，不含厂家/准字/规格关键词）
    if not info["药名"]:
        skip_kw = {"国药准字", "规格", "生产企业", "生产厂家", "厂家", "制药", "药业", "有限公司", "股份", "集团", "科技", "生物", "批准文号", "适应症", "用法用量", "贮藏", "有效期", "注意事项", "不良反应", "禁忌", "成份", "主要成份", "功能主治", "用法", "用量", "OTC", "处方药"}
        for ln in lines:
            clean = re.sub(r"[^\u4e00-\u9fffA-Za-z0-9]", "", ln)
            if 2 <= len(clean) <= 15 and not any(kw in ln for kw in skip_kw):
                # 再确认不是厂家/准字
                if clean not in info.get("厂家", "") and clean not in info.get("国药准字", ""):
                    info["药名"] = clean
                    break

    # 4.3 兜底：取第一个长度>=2的非关键词行
    if not info["药名"]:
        for ln in lines:
            clean = re.sub(r"[^\u4e00-\u9fffA-Za-z0-9]", "", ln)
            if len(clean) >= 2 and not any(kw in ln for kw in ["国药准字", "规格", "生产", "厂家", "制药", "药业", "有限公司"]):
                if clean not in info.get("厂家", "") and clean not in info.get("国药准字", ""):
                    info["药名"] = clean
                    break

    return info


# ===== OpenAI 兼容 API 视觉识别 =====
def _encode_image(image_path: str) -> str:
    """将图片编码为 base64"""
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")

def _get_image_mime(image_path: str) -> str:
    ext = os.path.splitext(image_path)[1].lower()
    mime_map = {
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".png": "image/png", ".webp": "image/webp",
        ".bmp": "image/bmp",
    }
    return mime_map.get(ext, "image/jpeg")

SYSTEM_PROMPT = """你是一个药品信息识别专家。请从药品包装/发票图片中提取药品信息，严格只返回 JSON 数组，不允许任何额外文字、注释、Markdown 代码块。

JSON 格式（每个对象必须包含以下 4 字段，缺失字段填空字符串""）：
[
  {
    "药名": "药品通用名称，如阿莫西林胶囊",
    "厂家": "生产企业完整名称",
    "国药准字": "国药准字+字母+8位数字，如国药准字H12345678",
    "规格": "如 0.25g*24粒、10ml*6支"
  }
]

注意：
- 如果图片中看不到某字段，该字段填 ""，严禁编造
- 只提取图片中实际可见的文字信息
- 多个药品按数组顺序返回
- 不要使用 ```json ... ``` 包裹，直接返回裸 JSON 数组
"""


def analyze_image(image_path: str, api_config: dict = None, model_dir: str = None) -> list[dict]:
    """
    统一分析接口
    - 如果有 api_config 且 api_key 不为空，使用 Vision API
    - 否则使用本地 EasyOCR
    """
    if api_config and api_config.get("api_key"):
        return extract_with_vision_api(image_path, api_config)
    else:
        text = extract_text_easyocr(image_path, model_dir)
        info = parse_medicine_info(text)
        return [info]

def extract_with_vision_api(image_path: str, api_config: dict) -> list[dict]:
    """使用 OpenAI 兼容的 Vision API 识别药品信息"""
    from openai import OpenAI
    import json, re

    base_url = api_config.get("base_url", "https://api.openai.com/v1")
    api_key = api_config.get("api_key")
    model = api_config.get("model", "gpt-4o")

    base64_image = _encode_image(image_path)
    mime_type = _get_image_mime(image_path)
    data_url = f"data:{mime_type};base64,{base64_image}"

    client = OpenAI(
        base_url=base_url,
        api_key=api_key,
    )

    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": USER_PROMPT},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            },
        ],
        max_tokens=2000,
        temperature=0.1,
    )

    # 诊断响应结构
    try:
        n_choices = len(response.choices)
        msg = response.choices[0].message if n_choices else None
        rc = getattr(msg, "reasoning_content", None) if msg else None
        role_name = getattr(msg, "role", None)
        content_type = type(getattr(msg, "content", None)).__name__
        print(f"[AI视觉] 响应结构: choices={n_choices} role={role_name} content类型={content_type}", flush=True)
        if msg and msg.content is None:
            print(f"[AI视觉] 警告: content 为 None（模型未返回文本）。reasoning_content存在={rc is not None}", flush=True)
    except Exception as diag_e:
        print(f"[AI视觉] 诊断异常: {diag_e}", flush=True)

    content_text = (response.choices[0].message.content or "").strip()
    print(f"[AI视觉] API返回({len(content_text)}字符): {content_text[:300]}", flush=True)

    # content 为空时，尝试用 reasoning_content（部分思考模型把答案放在这里）
    if not content_text:
        rc = getattr(response.choices[0].message, "reasoning_content", None)
        if rc:
            content_text = rc.strip()
            print(f"[AI视觉] 改用 reasoning_content 作为内容({len(content_text)}字符)", flush=True)

    # 严格解析 JSON：模型必须返回裸 JSON 数组
    content_text = re.sub(r"^```(?:json)?\s*|\s*```$", "", content_text, flags=re.MULTILINE).strip()
    
    try:
        result = json.loads(content_text)
        if isinstance(result, dict):
            result = [result]
        
        # 校验每个元素必须是 dict 且包含所有字段
        for item in result:
            if not isinstance(item, dict):
                raise ValueError("Each item must be an object")
            for field in MEDICINE_FIELDS:
                if field not in item:
                    item[field] = ""
        return result
    except (json.JSONDecodeError, ValueError) as e:
        print(f"[AI视觉] JSON 解析失败: {e}", flush=True)
        print(f"[AI视觉] 原始内容: {content_text[:500]}", flush=True)
        raise RuntimeError(f"AI 返回非标准 JSON，无法解析: {e}\n原始返回: {content_text[:500]}")




import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def export_to_excel(results, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '药品识别结果'
    headers = MEDICINE_FIELDS + ['_warn']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for row_idx, item in enumerate(results, 2):
        for col_idx, field in enumerate(headers, 1):
            value = item.get(field, '')
            if field == '_warn' and isinstance(value, list):
                value = '; '.join(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True)
            if field == '_warn' and value:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=c).font = Font(color='FF0000')
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    wb.save(file_path)

